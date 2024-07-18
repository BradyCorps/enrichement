import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import io
import json
import os

# Define global variables for storing SKU and SEQ/NAME data
sku_data_list = []
seq_name_data_list = []
taxonomy_groups = {}
skip_taxonomies = set()  # To keep track of skipped taxonomies

# File to store past runs
history_file = 'enrichment_history.json'

def save_history():
    history = {
        "runs": []
    }
    if os.path.exists(history_file):
        with open(history_file, 'r') as file:
            history = json.load(file)
    
    run_data = {
        "sku_data": sku_data_list,
        "seq_name_data": seq_name_data_list
    }
    
    history["runs"].insert(0, run_data)
    history["runs"] = history["runs"][:3]  # Keep only last 3 runs
    
    with open(history_file, 'w') as file:
        json.dump(history, file)

def create_excel():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Enrichment"
        
        # Add SKU headers
        sku_headers = [
            'SKU #', 'SKU Title (English)', 'Attribute value of PKG Product Identity Modifier (English, DEFAULT)', 
            'Attribute value of PKG Product Identity (English, DEFAULT)', 'Attribute value of PKG Custom Callout (English, DEFAULT)', 
            'Attribute value of Ideal for (English, DEFAULT)', 'Structure group(s) (Brands Structure)', 'PKG Current Package Type', 
            'Structure assignments (Selling Taxonomy)', 'Structure assignments (Promotion Structure)', 'MMS Item Status', 'Vendor'
        ]
        ws.append(sku_headers)
        
        # Highlight the header row
        fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for cell in ws[1]:
            cell.fill = fill_yellow
        
        # Add SKU data rows starting at row 2
        for sku_data in sku_data_list:
            sku_df = pd.read_csv(io.StringIO(sku_data), delimiter='\t')
            for _, row in sku_df.iterrows():
                ws.append(row.tolist())
                # Extract Structure assignments and initialize combined data
                taxonomy = row['Structure assignments (Selling Taxonomy)']
                if taxonomy not in taxonomy_groups:
                    taxonomy_groups[taxonomy] = []
        
        # Add secondary headers
        ws.append([])
        secondary_headers = ['Packaging', 'Selling', 'Warning', 'Concerns']
        ws.append([""] + secondary_headers)
        
        # Add green row after secondary headers
        fill_light_green = PatternFill(start_color="D0F0C0", end_color="D0F0C0", fill_type="solid")
        for cell in ws[ws.max_row]:
            cell.fill = fill_light_green
        
        # Write SEQ and NAME data to Excel directly after each Step 2
        for seq_name_data in seq_name_data_list:
            seq_name_df = pd.read_csv(io.StringIO(seq_name_data), delimiter='\t')
            for _, row in seq_name_df.iterrows():
                name = row.get('Name (English)', '')
                attribute_value = row.get('Attribute value (English, DEFAULT)', '')
                purpose = row.get('Purpose', '')
                ws.append([""] + [name, attribute_value, purpose])
            # Add green row after each set of SEQ and NAME data
            ws.append([])  # Insert empty row
            for cell in ws[ws.max_row]:
                cell.fill = fill_light_green
        
        # Save the workbook
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filename:
            wb.save(filename)
            save_history()
            messagebox.showinfo("Success", f"File saved as {filename}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while creating the Excel file: {str(e)}")

def paste_sku_data():
    try:
        sku_data = text_area.get("1.0", tk.END).strip()
        if not sku_data:
            raise ValueError("SKU data is empty.")
        sku_data_list.append(sku_data)
        text_area.delete("1.0", tk.END)
        step_2_button.config(state=tk.NORMAL)
        back_button.config(state=tk.NORMAL)
        messagebox.showinfo("Success", "Step 1 completed. Now proceed to Step 2.")
    except Exception as e:
        messagebox.showerror("Error", f"Step 1 Error: {str(e)}")

def paste_seq_name_data():
    try:
        seq_name_data = text_area.get("1.0", tk.END).strip()
        if not seq_name_data:
            raise ValueError("SEQ and NAME data is empty.")
        seq_name_data_list.append(seq_name_data)
        text_area.delete("1.0", tk.END)
        add_another_sku_button.config(state=tk.NORMAL)
        complete_button.config(state=tk.NORMAL)
        skip_button.config(state=tk.NORMAL)  # Enable skip button
        messagebox.showinfo("Success", "Step 2 completed. You can add another SKU, skip this data, or complete the process.")
    except Exception as e:
        messagebox.showerror("Error", f"Step 2 Error: {str(e)}")

def add_another_sku():
    step_2_button.config(state=tk.DISABLED)
    messagebox.showinfo("Info", "Ready for another SKU. Please paste SKU data for the next product.")

def skip_step_2():
    global skip_taxonomies
    # Optionally, you can prompt the user for which taxonomy to skip
    skip_taxonomies = {taxonomy for taxonomy in taxonomy_groups.keys()}
    messagebox.showinfo("Info", "All data for Step 2 has been skipped for the current SKU.")

def complete_enrichment():
    create_excel()

def recall_last_run(run_index):
    try:
        with open(history_file, 'r') as file:
            history = json.load(file)
        run_data = history["runs"][run_index]
        global sku_data_list, seq_name_data_list
        sku_data_list = run_data["sku_data"]
        seq_name_data_list = run_data["seq_name_data"]
        # Populate the GUI with SKU numbers
        sku_numbers = [pd.read_csv(io.StringIO(sku_data), delimiter='\t')['SKU #'].iloc[0] for sku_data in sku_data_list]
        text_area.insert(tk.END, "\n".join(map(str, sku_numbers)))
        messagebox.showinfo("Success", f"Recalled run {run_index + 1}")
    except Exception as e:
        messagebox.showerror("Error", f"Error recalling run {run_index + 1}: {str(e)}")

def go_back():
    if seq_name_data_list:
        seq_name_data_list.pop()
        complete_button.config(state=tk.DISABLED)
        add_another_sku_button.config(state=tk.DISABLED)
        skip_button.config(state=tk.DISABLED)  # Disable skip button
        step_2_button.config(state=tk.NORMAL)
        messagebox.showinfo("Info", "Went back to Step 2.")
    elif sku_data_list:
        sku_data_list.pop()
        step_2_button.config(state=tk.DISABLED)
        back_button.config(state=tk.DISABLED)
        messagebox.showinfo("Info", "Went back to Step 1.")

def clear_data():
    global sku_data_list, seq_name_data_list, taxonomy_groups, skip_taxonomies
    sku_data_list = []
    seq_name_data_list = []
    taxonomy_groups = {}
    skip_taxonomies = set()
    text_area.delete("1.0", tk.END)
    step_2_button.config(state=tk.DISABLED)
    back_button.config(state=tk.DISABLED)
    add_another_sku_button.config(state=tk.DISABLED)
    complete_button.config(state=tk.DISABLED)
    skip_button.config(state=tk.DISABLED)
    messagebox.showinfo("Info", "All data cleared. Ready to start fresh.")

# Create the main window
root = tk.Tk()
root.title("Data Formatter")

# Create and place the text area for raw data input
text_area = tk.Text(root, wrap='word', width=100, height=20)
text_area.pack(padx=10, pady=10)

# Create and place the buttons for each step
step_1_button = tk.Button(root, text="Step 1: Paste SKU Data", command=paste_sku_data)
step_1_button.pack(pady=10)

step_2_button = tk.Button(root, text="Step 2: Paste SEQ and NAME Data", command=paste_seq_name_data, state=tk.DISABLED)
step_2_button.pack(pady=10)

add_another_sku_button = tk.Button(root, text="Add Another SKU", command=add_another_sku, state=tk.DISABLED)
add_another_sku_button.pack(pady=10)

skip_button = tk.Button(root, text="Skip Step 2 for Current SKU", command=skip_step_2, state=tk.DISABLED)
skip_button.pack(pady=10)

complete_button = tk.Button(root, text="Complete and Generate Excel", command=complete_enrichment, state=tk.DISABLED)
complete_button.pack(pady=10)

back_button = tk.Button(root, text="Go Back to Previous SKU", command=go_back, state=tk.DISABLED)
back_button.pack(pady=10)

clear_button = tk.Button(root, text="Clear All Data", command=clear_data)
clear_button.pack(pady=10)

recall_last_button = tk.Button(root, text="Recall Last Run", command=lambda: recall_last_run(0))
recall_last_button.pack(pady=10)

# Start the main loop
root.mainloop()
